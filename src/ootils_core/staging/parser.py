"""
⚠️  DEPRECATED (ADR-042, 2026-07-18) — the staging PIPELINE is dead
(`api/routers/staging.py` is unmounted, see its banner; the pipeline never
reached `status='validated'` in production and is superseded by the governed
daily-run pipeline). This PARSER remains directly imported (not via HTTP) by
`tests/test_staging_parser.py` and `tests/test_ingest_status_lifecycle.py` to
exercise the TSV/CSV/XLSX/JSON parsing logic in isolation. Module + tests are
kept, not dropped; do not wire this back behind a live endpoint.

parser.py — unified file parser for the staging pipeline (ADR-013 D1).

Accepts file bytes + an optional format hint, returns a `ParseResult`
with the parsed rows as dicts keyed by the header columns, plus the
metadata needed by `staging.uploads` (format, encoding, sha256).

Four supported formats:

  TSV   .tab / .tsv         delimiter '\\t'. Recommended for ERP exports
                            because business values rarely contain tabs.
  CSV   .csv                delimiter auto-detected via csv.Sniffer
                            (',' or ';'). Quoting '"' standard.
  XLSX  .xlsx               first sheet by default; pass sheet_name in
                            ParseOptions to override. openpyxl read-only.
  JSON  .json               top-level array of objects, one per row.

All formats converge to the same intermediate shape: a list of
``dict[str, str]`` where keys are the column headers (case-preserved
from the file) and values are TEXT (no type coercion happens here —
that's the DQ engine's job downstream).

Encoding handling for text formats:
  1. Try UTF-8 (with or without BOM)
  2. Fall back to CP-1252 (typical for legacy Windows / SAP exports),
     log a warning
  3. Anything else: raise ParseError

Format detection (when not given):
  - by extension if present and recognised
  - else content-sniffed (PK -> xlsx, { or [ -> json, \\t in header -> tsv)
  - else CSV by default
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)


SUPPORTED_FORMATS = ("tsv", "csv", "xlsx", "json")
TEXT_FORMATS = ("tsv", "csv", "json")


# Encodings we try in order, in priority. Add to this tuple if a new
# legacy source needs support; the first one that decodes wins.
_ENCODINGS_TRY = ("utf-8-sig", "utf-8", "cp1252", "latin-1")


class ParseError(ValueError):
    """Raised when the file cannot be parsed (format, encoding, structure)."""


@dataclass(frozen=True)
class ParseOptions:
    """Tunables for the parser. Defaults match the most common ERP shape.

    sheet_name: only relevant for XLSX. None -> first sheet.
    delimiter:  if set, skip csv.Sniffer and use this delimiter directly.
                Use when the source is known (e.g. always ';' from a
                French ERP).
    max_rows:   if set, stop parsing after this many data rows (excluding
                the header). Useful for dry-run / preview endpoints.
    """
    sheet_name: str | None = None
    delimiter: str | None = None
    max_rows: int | None = None


@dataclass
class ParseResult:
    """Outcome of parsing one file. All fields are populated except
    when a format doesn't have a notion of the metadata (e.g. XLSX
    has no `encoding`, JSON has no `delimiter`)."""
    rows: list[dict[str, str]]
    headers: list[str]
    format: str               # 'tsv' / 'csv' / 'xlsx' / 'json'
    encoding: str | None      # text formats only
    delimiter: str | None     # tsv / csv only
    sha256: str
    # Diagnostic counters for the response
    row_count: int = field(init=False)
    header_count: int = field(init=False)

    def __post_init__(self) -> None:
        # Use object.__setattr__ since dataclass is mutable but these
        # are derived fields — we don't want callers messing with them.
        object.__setattr__(self, "row_count", len(self.rows))
        object.__setattr__(self, "header_count", len(self.headers))


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def parse(
    data: bytes,
    filename: str | None = None,
    format_hint: str | None = None,
    options: ParseOptions | None = None,
) -> ParseResult:
    """Parse `data` (the raw file bytes) into a ParseResult.

    Arguments:
        data:         file bytes, exactly as read from the upload
        filename:     optional original filename. Used for format
                      detection by extension. Not stored elsewhere.
        format_hint:  optional explicit format. If set, skips detection.
                      Must be one of SUPPORTED_FORMATS.
        options:      optional ParseOptions for advanced control.

    Returns:
        ParseResult with rows + metadata.

    Raises:
        ParseError if the file cannot be decoded, sniffed, or parsed.
    """
    if not data:
        raise ParseError("empty file")

    opts = options or ParseOptions()
    fmt = (format_hint or _detect_format(data, filename) or "csv").lower()
    if fmt not in SUPPORTED_FORMATS:
        raise ParseError(f"unsupported format: {fmt!r}")

    sha = hashlib.sha256(data).hexdigest()

    if fmt == "xlsx":
        rows, headers = _parse_xlsx(data, opts)
        return ParseResult(
            rows=rows, headers=headers, format="xlsx",
            encoding=None, delimiter=None, sha256=sha,
        )
    if fmt == "json":
        text, enc = _decode_text(data)
        rows, headers = _parse_json(text)
        return ParseResult(
            rows=rows, headers=headers, format="json",
            encoding=enc, delimiter=None, sha256=sha,
        )
    # TSV / CSV
    text, enc = _decode_text(data)
    delim = opts.delimiter or _sniff_delimiter(text, fmt)
    rows, headers = _parse_delimited(text, delim, opts.max_rows, fmt=fmt)
    return ParseResult(
        rows=rows, headers=headers, format=fmt,
        encoding=enc, delimiter=delim, sha256=sha,
    )


# ---------------------------------------------------------------------------
# Format detection
# ---------------------------------------------------------------------------


_EXTENSION_MAP = {
    ".tsv":  "tsv",
    ".tab":  "tsv",
    ".csv":  "csv",
    ".xlsx": "xlsx",
    ".json": "json",
}


def _detect_format(data: bytes, filename: str | None) -> str | None:
    """Best-effort format detection. Returns None if undecidable."""
    if filename:
        # Find the last dot-extension (lowercase)
        idx = filename.rfind(".")
        if idx >= 0:
            ext = filename[idx:].lower()
            if ext in _EXTENSION_MAP:
                return _EXTENSION_MAP[ext]

    # Content sniffing
    head = data[:64]
    # XLSX is a ZIP container — starts with "PK"
    if head.startswith(b"PK\x03\x04") or head.startswith(b"PK\x05\x06"):
        return "xlsx"
    # JSON: array or object as outermost (skip whitespace)
    stripped = head.lstrip()
    if stripped.startswith(b"[") or stripped.startswith(b"{"):
        return "json"
    # Try to decode the first line and check for a tab
    try:
        first_nl = data.find(b"\n")
        first_line_bytes = data[: first_nl if first_nl > 0 else 200]
        first_line = first_line_bytes.decode("utf-8", errors="replace")
    except Exception:
        return None
    if "\t" in first_line:
        return "tsv"
    return "csv"


# ---------------------------------------------------------------------------
# Text decoding (encoding detection)
# ---------------------------------------------------------------------------


def _decode_text(data: bytes) -> tuple[str, str]:
    """Try the encoding ladder. Returns (text, reported_encoding).

    The codec `utf-8-sig` is used first because it strips the BOM when
    present and behaves like plain `utf-8` otherwise. We report
    `utf-8-sig` only when the file actually had a BOM, otherwise we
    report `utf-8` — the distinction matters for downstream diagnostics
    (a BOM in an ERP export is a signal worth noting in the response).
    """
    had_bom = data.startswith(b"\xef\xbb\xbf")
    for enc in _ENCODINGS_TRY:
        try:
            text = data.decode(enc)
            if enc in ("utf-8", "utf-8-sig"):
                reported = "utf-8-sig" if had_bom else "utf-8"
            else:
                reported = enc
                logger.warning(
                    "staging.parser: file decoded with %s (non-UTF-8). "
                    "Consider asking the source system to export UTF-8.",
                    enc,
                )
            return text, reported
        except UnicodeDecodeError:
            continue
    raise ParseError(
        f"could not decode file with any of {list(_ENCODINGS_TRY)}; "
        "check the source-system export settings"
    )


# ---------------------------------------------------------------------------
# Delimiter sniffing (TSV / CSV)
# ---------------------------------------------------------------------------


def _sniff_delimiter(text: str, fmt: str) -> str:
    """For TSV, always '\\t'. For CSV, try csv.Sniffer on the first line."""
    if fmt == "tsv":
        return "\t"
    # CSV — try Sniffer; fall back to ',' if it fails
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        return ","


# ---------------------------------------------------------------------------
# Per-format parsers
# ---------------------------------------------------------------------------


def _parse_delimited(
    text: str, delimiter: str, max_rows: int | None, fmt: str = "csv"
) -> tuple[list[dict[str, str]], list[str]]:
    """Parse a delimited text blob into (rows, headers).

    The first non-empty line is the header. Trailing whitespace on each
    cell is stripped. Empty cells become "" (not None) — typing happens
    in the DQ pipeline, not here.

    Quoting: CSV keeps standard `"`-quoting (module docstring, §CSV) since
    a comma-delimited value legitimately needs it. TSV disables quoting
    entirely (`csv.QUOTE_NONE`) — same rationale as `scripts/ingest_file.py`
    and TSV-FILES-SPEC.md §1.1 ("aucun guillemet"): a literal `"` in a cell
    (e.g. an inch-mark item description) must be preserved verbatim, never
    interpreted as a CSV quote character.
    """
    buf = io.StringIO(text)
    if fmt == "tsv":
        reader = csv.reader(buf, delimiter=delimiter, quoting=csv.QUOTE_NONE)
    else:
        reader = csv.reader(buf, delimiter=delimiter, quotechar='"')

    headers: list[str] = []
    rows: list[dict[str, str]] = []

    for line_no, raw in enumerate(reader, start=1):
        if not raw or all(cell.strip() == "" for cell in raw):
            continue  # skip blank lines
        if not headers:
            headers = [c.strip() for c in raw]
            if not all(headers):
                raise ParseError(
                    f"header line contains an empty column at index "
                    f"{headers.index('')} (line {line_no})"
                )
            if len(set(h.lower() for h in headers)) != len(headers):
                raise ParseError(
                    f"header line contains duplicate columns "
                    f"(case-insensitive): {headers}"
                )
            continue
        # Pad / truncate to header length so the dict construction works
        cells = [c.strip() for c in raw]
        if len(cells) < len(headers):
            cells = cells + [""] * (len(headers) - len(cells))
        elif len(cells) > len(headers):
            cells = cells[: len(headers)]
        rows.append(dict(zip(headers, cells)))
        if max_rows is not None and len(rows) >= max_rows:
            break

    if not headers:
        raise ParseError("file contains no header line")
    return rows, headers


def _parse_json(text: str) -> tuple[list[dict[str, str]], list[str]]:
    """Parse a JSON array of objects. All scalar values stringified."""
    try:
        obj = json.loads(text)
    except json.JSONDecodeError as e:
        raise ParseError(f"invalid JSON: {e}") from e
    if not isinstance(obj, list):
        raise ParseError(
            f"JSON top-level must be an array of objects, got {type(obj).__name__}"
        )
    if not obj:
        return [], []
    if not isinstance(obj[0], dict):
        raise ParseError(
            f"JSON array elements must be objects, got {type(obj[0]).__name__}"
        )
    # Headers = union of all keys in the array, ordered by first appearance.
    headers: list[str] = []
    seen: set[str] = set()
    for item in obj:
        if not isinstance(item, dict):
            continue
        for k in item.keys():
            if k not in seen:
                seen.add(k)
                headers.append(k)

    rows: list[dict[str, str]] = []
    for item in obj:
        if not isinstance(item, dict):
            continue
        row = {h: _scalar_to_str(item.get(h)) for h in headers}
        rows.append(row)
    return rows, headers


def _parse_xlsx(data: bytes, options: ParseOptions) -> tuple[list[dict[str, str]], list[str]]:
    """Parse the chosen sheet. Empty cells become '' (not None)."""
    try:
        # Lazy import — openpyxl is a heavy dependency we only load if needed
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ParseError(
            "openpyxl is required to parse XLSX files but is not installed"
        ) from exc

    try:
        wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise ParseError(f"invalid XLSX: {e}") from e

    if options.sheet_name:
        if options.sheet_name not in wb.sheetnames:
            raise ParseError(
                f"sheet {options.sheet_name!r} not found; "
                f"available: {wb.sheetnames}"
            )
        ws = wb[options.sheet_name]
    else:
        ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    # First non-empty row = header
    headers: list[str] = []
    for raw in rows_iter:
        if not raw or all(v is None or str(v).strip() == "" for v in raw):
            continue
        headers = [str(v).strip() if v is not None else "" for v in raw]
        if not all(headers):
            raise ParseError("header row contains an empty column")
        break

    if not headers:
        raise ParseError("XLSX contains no header row")

    rows: list[dict[str, str]] = []
    max_rows = options.max_rows
    for raw in rows_iter:
        if not raw or all(v is None or str(v).strip() == "" for v in raw):
            continue
        cells: list[str] = []
        for v in raw[: len(headers)]:
            cells.append(_scalar_to_str(v))
        if len(cells) < len(headers):
            cells = cells + [""] * (len(headers) - len(cells))
        rows.append(dict(zip(headers, cells)))
        if max_rows is not None and len(rows) >= max_rows:
            break

    return rows, headers


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _scalar_to_str(v) -> str:
    """Convert any scalar (None, str, int, float, bool, date) to a TEXT-safe string.

    The staging layer is type-agnostic by design; coercion + validation
    happens in the DQ engine. So we just produce a stable text form here.
    """
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, str):
        return v.strip()
    # date, datetime, Decimal, int, float — str() gives a canonical form
    return str(v).strip()
